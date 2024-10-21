import os

import geopandas as gpd
import matplotlib.colors as mcolors
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import rasterio
import rasterio.features
from affine import Affine
from openpyxl import Workbook
from rasterio import features
from rasterio.enums import Resampling
from rasterio.features import rasterize
from rasterio.plot import show
from rasterio.windows import from_bounds
from scipy.ndimage import binary_dilation, binary_erosion

# from midar.ign.utils import get_window
from shapely.geometry import box, mapping, shape
from tqdm import tqdm


def get_mask(
    image_path, min_area=1000000, plot=True, plot_path="mask.jpg", crs=2154
):
    with rasterio.open(image_path) as src:
        image = src.read(1)
        mask = image == 0
        polygons = [
            shape(geom)
            for geom, value in tqdm(
                features.shapes(
                    mask.astype(np.int16), transform=src.transform
                ),
                desc="Retrieving mask",
            )
            if value == 1
        ]

    # Convert polygons to GeoDataFrame
    gdf = gpd.GeoDataFrame({"geometry": polygons}, crs=crs)

    # FIlter with min_area
    gdf = gdf[gdf.geometry.area > min_area]

    if plot:
        vmin = 0
        vmax = 40

        fig, axes = plt.subplots(1, 1, figsize=(10, 10))
        # Labels
        show(
            image,
            ax=axes,
            transform=src.transform,
            vmin=vmin,
            vmax=vmax,
            cmap="Greens",
        )
        gdf.boundary.plot(ax=axes, edgecolor="red", linewidth=2)
        fig.savefig(plot_path, bbox_inches="tight", pad_inches=0)

    return gdf


def mask_with_geometry(tif_path, geometry, output_tif_path, nodata=np.nan):
    # Load the TIFF file
    with rasterio.open(tif_path) as src:
        tif_data = src.read(1)  # Read the first band
        tif_meta = src.meta.copy()

        # Create a mask based on the geometry
        mask = rasterio.features.geometry_mask(
            [mapping(geometry)],
            transform=src.transform,
            invert=False,
            out_shape=src.shape,
        )
        # Apply the mask to the TIFF data
        masked_tif_data = np.where(
            mask, tif_data, nodata
        )  # Using NaN to indicate masked areas
        # Update metadata to ensure compatibility
        if pd.isnull(nodata):
            tif_meta.update(dtype="float32")
        # Save the masked output to a new TIFF file
        with rasterio.open(output_tif_path, "w", **tif_meta) as dest:
            dest.write(masked_tif_data.astype(np.float32), 1)


def get_changes(
    image_1_path,
    image_2_path,
    delta,
    min_area=5000,
    bounds=None,
    scaling_factor_1=None,
    scaling_factor_2=None,
    classification_mask_path=None,
    classes_to_keep=[1, 2, 3, 4, 5, 6],
    resolution=None,
):
    # Assuming both images have the same profile
    image_1, profile = get_window(
        image_1_path,
        geometry=None,
        bounds=bounds,
        resolution=resolution,
        return_profile=True,
        resampling_method="bilinear",
    )
    if (
        image_2_path is None
    ):  # handle special case when only loss is available and given as image_1
        image_2 = np.zeros_like(image_1)
    else:
        image_2 = get_window(
            image_2_path,
            geometry=None,
            bounds=bounds,
            resolution=resolution,
            resampling_method="bilinear",
        )

    if scaling_factor_1:
        image_1 = image_1 * scaling_factor_1
    if scaling_factor_2:
        image_2 = image_2 * scaling_factor_2

    image_1 = image_1.squeeze()
    image_2 = image_2.squeeze()

    if classification_mask_path:
        mask = get_window(
            classification_mask_path,
            geometry=None,
            bounds=bounds,
            resolution=resolution,
            return_profile=False,
            resampling_method="bilinear",
        ).squeeze()
        for ix in np.unique(mask):
            if ix not in classes_to_keep:
                image_1[mask == ix] = 0
                image_2[mask == ix] = 0

    difference = image_2.astype(np.float32) - image_1.astype(np.float32)
    changes = difference < delta

    size = 3
    # Apply morphology to filter out noise
    print("Apply morphology")
    changes = binary_erosion(changes, structure=np.ones((size, size)))
    changes = binary_dilation(
        changes, structure=np.ones((size, size)), iterations=2
    )
    changes = binary_erosion(changes, structure=np.ones((size, size)))

    # Transform array into polygons
    polygons = [
        shape(geom)
        for geom, value in features.shapes(
            changes.astype(np.int16), transform=profile["transform"]
        )
        if value == 1
    ]

    # Convert polygons to GeoDataFrame
    gdf = gpd.GeoDataFrame({"geometry": polygons}, crs=profile["crs"])

    # Filter with min_area (not necessary since already filtered)
    gdf_filtered = gdf[gdf.geometry.area > min_area]

    # Mask the changes
    # Create a mask based on the geometry
    if len(gdf_filtered["geometry"].values):
        changes = rasterio.features.geometry_mask(
            gdf_filtered["geometry"].values,
            transform=profile["transform"],
            invert=True,
            out_shape=changes.shape,
        )
    else:
        changes = np.zeros(changes.shape, dtype=bool)

    return image_1, image_2, difference, changes, gdf_filtered


def compute_iou(mask1, mask2):
    intersection = np.logical_and(mask1, mask2)
    union = np.logical_or(mask1, mask2)
    iou = np.sum(intersection) / np.sum(union)
    return iou


def compute_tree_cover(tif_path, threshold, scaling_factor=None):
    with rasterio.open(tif_path) as src:
        canopy_height = src.read(1)
        if scaling_factor:
            canopy_height = canopy_height * scaling_factor
        # Create a binary mask where canopy height is greater than the threshold
        tree_cover = canopy_height > threshold
    return tree_cover


def compute_change_metrics(
    mask1, mask2, difference_1, difference_2, resolution=1.5
):
    tp = np.sum(np.logical_and(mask1, mask2))  # True positives
    # tn = np.sum(np.logical_and(~mask1, ~mask2))  # True negatives
    fp = np.sum(
        np.logical_and(mask2, np.logical_not(mask1))
    )  # False positives
    fn = np.sum(
        np.logical_and(mask1, np.logical_not(mask2))
    )  # False negatives

    res = {}
    res["precision"] = tp / (tp + fp) if tp + fp > 0 else 0
    res["recall"] = tp / (tp + fn) if tp + fn > 0 else 0
    res["f1"] = (
        2
        * (res["precision"] * res["recall"])
        / (res["precision"] + res["recall"])
        if res["precision"] + res["recall"] > 0
        else 0
    )
    res["iou"] = compute_iou(mask1, mask2)

    return res


def save_dict_of_dicts_to_excel(dict_of_dicts, filename):
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    row = 1
    for table_name, sub_dict in dict_of_dicts.items():
        # Write the name of the dictionary in the top-left cell
        ws.cell(row=row, column=1, value=table_name)

        # Write keys in the first row after the table name
        for col, key in enumerate(sub_dict.keys(), start=1):
            ws.cell(row=row + 1, column=col, value=key)

        # Write values in the second row after the keys
        for col, value in enumerate(sub_dict.values(), start=1):
            ws.cell(row=row + 2, column=col, value=value)

        # Move to the next position, leaving a blank row
        row += 5

    # Save the workbook to the specified file
    wb.save(filename)


def plot_detections(
    inputs_1_path,
    inputs_2_path,
    bounds,
    metrics_out,
    fs=16,
    save_prefix="",
    year_1=2022,
    year_2=2023,
    save_dir="./",
):
    # Open one of the images to get the spatial context (bounds and CRS)
    fig, axes = plt.subplots(3, 3, figsize=(10, 10))

    with rasterio.open(inputs_1_path) as src1, rasterio.open(
        inputs_2_path
    ) as src2:
        window = src1.window(*bounds)
        image_1 = src1.read(window=window)[:3]
        window = src2.window(*bounds)
        image_2 = src2.read(window=window)[:3]

    def clip_image(image, max_value=0.3):
        image = image / 255
        image = image.clip(min=0, max=max_value)
        image = image / image.max()
        return image

    image_1 = clip_image(image_1)
    image_2 = clip_image(image_2)

    # Plot the images

    # Create the custom colormap for change
    colors = [(0, "red"), (0.5, "yellow"), (1, "green")]
    cmap_diff = mcolors.LinearSegmentedColormap.from_list(
        "custom_cmap", colors
    )
    norm = mcolors.TwoSlopeNorm(vmin=-20, vcenter=0, vmax=2)

    vmin = 0
    vmax = 40

    # Inputs
    axes[0, 0].imshow(image_1.transpose(1, 2, 0))
    axes[0, 0].set_title(f"Spot {year_1}")
    axes[0, 0].axis("off")  # Turn off the axis

    axes[0, 1].imshow(image_2.transpose(1, 2, 0))
    axes[0, 1].set_title(f"Spot {year_2}")
    axes[0, 1].axis("off")  # Turn off the axis

    axes[0, 2].imshow(metrics_out["detection_lidar"], cmap="inferno")
    axes[0, 2].set_title("Lidar detected changes")
    axes[0, 2].axis("off")  # Turn off the axis

    # Labels
    axes[1, 0].imshow(
        metrics_out["height_1"], vmin=vmin, vmax=vmax, cmap="Greens"
    )
    axes[1, 0].set_title(f"Lidar {year_1}")
    axes[1, 0].axis("off")  # Turn off the axis

    axes[1, 1].imshow(
        metrics_out["height_2"], vmin=vmin, vmax=vmax, cmap="Greens"
    )
    axes[1, 1].set_title(f"Lidar {year_2}")
    axes[1, 1].axis("off")  # Turn off the axis

    axes[1, 2].imshow(metrics_out["difference"], cmap=cmap_diff, norm=norm)
    # axes[1, 2].imshow(-metrics_out["difference"], vmin=vmin_diff, vmax=vmax_diff, cmap="inferno")
    axes[1, 2].set_title("Lidar Height Difference")
    axes[1, 2].axis("off")  # Turn off the axis

    # Predictions
    axes[2, 0].imshow(
        metrics_out["pred_1"], vmin=vmin, vmax=vmax, cmap="Greens"
    )
    axes[2, 0].set_title(f"Predictions {year_1}")
    axes[2, 0].axis("off")  # Turn off the axis

    axes[2, 1].imshow(
        metrics_out["pred_2"], vmin=vmin, vmax=vmax, cmap="Greens"
    )
    axes[2, 1].set_title(f"Predictions {year_2}")
    axes[2, 1].axis("off")  # Turn off the axis

    axes[2, 2].imshow(
        metrics_out["difference_pred"], cmap=cmap_diff, norm=norm
    )
    # axes[2, 2].imshow(-metrics_out["difference_pred"], vmin=vmin_diff, vmax=vmax_diff, cmap="inferno")
    axes[2, 2].set_title("Predictions Height Difference")
    axes[2, 2].axis("off")  # Turn off the axis

    fig.savefig(
        os.path.join(save_dir, f"{save_prefix}_detection_images.jpg"),
        bbox_inches="tight",
        pad_inches=0,
    )

    # Save also individual figures for paper
    save_dir_ind = os.path.join(save_dir, "individual_figures")
    if not os.path.exists(save_dir_ind):
        os.makedirs(save_dir_ind)

    fig, ax = plt.subplots()
    plt.imshow(image_1.transpose(1, 2, 0))
    plt.axis("off")  # Turn off the axis
    fig.savefig(
        os.path.join(save_dir_ind, f"{save_prefix}_spot_1.jpg"),
        bbox_inches="tight",
        pad_inches=0,
    )

    plt.imshow(image_2.transpose(1, 2, 0))
    plt.axis("off")  # Turn off the axis
    fig.savefig(
        os.path.join(save_dir_ind, f"{save_prefix}_spot_2.jpg"),
        bbox_inches="tight",
        pad_inches=0,
    )

    plt.imshow(metrics_out["detection_lidar"], cmap="inferno")
    plt.axis("off")  # Turn off the axis
    fig.savefig(
        os.path.join(save_dir_ind, f"{save_prefix}_detection_lidar.jpg"),
        bbox_inches="tight",
        pad_inches=0,
    )

    plt.imshow(metrics_out["detection_pred"], cmap="inferno")
    plt.axis("off")  # Turn off the axis
    fig.savefig(
        os.path.join(save_dir_ind, f"{save_prefix}_detection_pred.jpg"),
        bbox_inches="tight",
        pad_inches=0,
    )

    # Labels
    plt.imshow(metrics_out["height_1"], vmin=vmin, vmax=vmax, cmap="Greens")
    plt.axis("off")  # Turn off the axis
    fig.savefig(
        os.path.join(save_dir_ind, f"{save_prefix}_lidar_1.jpg"),
        bbox_inches="tight",
        pad_inches=0,
    )

    plt.imshow(metrics_out["height_2"], vmin=vmin, vmax=vmax, cmap="Greens")
    plt.axis("off")  # Turn off the axis
    fig.savefig(
        os.path.join(save_dir_ind, f"{save_prefix}_lidar_2.jpg"),
        bbox_inches="tight",
        pad_inches=0,
    )

    plt.imshow(metrics_out["difference"], cmap=cmap_diff, norm=norm)
    # plt.imshow(-metrics_out["difference"], vmin=vmin_diff, vmax=vmax_diff, cmap="inferno")
    plt.axis("off")  # Turn off the axis
    fig.savefig(
        os.path.join(save_dir_ind, f"{save_prefix}_lidar_diff.jpg"),
        bbox_inches="tight",
        pad_inches=0,
    )

    # Predictions
    plt.imshow(metrics_out["pred_1"], vmin=vmin, vmax=vmax, cmap="Greens")
    plt.axis("off")  # Turn off the axis
    fig.savefig(
        os.path.join(save_dir_ind, f"{save_prefix}_pred_1.jpg"),
        bbox_inches="tight",
        pad_inches=0,
    )

    plt.imshow(metrics_out["pred_2"], vmin=vmin, vmax=vmax, cmap="Greens")
    plt.axis("off")  # Turn off the axis
    fig.savefig(
        os.path.join(save_dir_ind, f"{save_prefix}_pred_2.jpg"),
        bbox_inches="tight",
        pad_inches=0,
    )

    plt.imshow(metrics_out["difference_pred"], cmap=cmap_diff, norm=norm)
    # plt.imshow(-metrics_out["difference_pred"], vmin=vmin_diff, vmax=vmax_diff, cmap="inferno")
    plt.axis("off")  # Turn off the axis
    fig.savefig(
        os.path.join(save_dir_ind, f"{save_prefix}_diff_pred.jpg"),
        bbox_inches="tight",
        pad_inches=0,
    )

    # Plot metrics
    metrics_list = ["precision", "recall", "f1", "iou"]
    fig = plt.figure(figsize=(10, 10))

    cmap = plt.cm.gray
    ax = plt.subplot2grid((2, 2), (0, 0))
    ax.imshow(metrics_out["detection_lidar"], cmap=cmap)
    ax.axis("off")  # Turn off the axis
    ax.set_title("Orig - Lidar (L)", fontsize=fs)
    ax = plt.subplot2grid((2, 2), (1, 0))
    ax.imshow(metrics_out["detection_pred"], cmap=cmap)
    ax.axis("off")  # Turn off the axis
    ax.set_title("Orig - Pred (P)", fontsize=fs)

    ax = plt.subplot2grid((2, 2), (0, 1), rowspan=2)
    ax.bar(metrics_list, [metrics_out["metrics"][key] for key in metrics_list])
    ax.set_title("Orig - scores L vs. P", fontsize=fs)
    ax.set_ylim(0, 1)
    ax.grid(color="gray", linestyle="dashed", axis="y")
    plt.xticks(rotation=40, fontsize=fs)
    plt.yticks(fontsize=fs)

    fig.savefig(
        os.path.join(save_dir, f"{save_prefix}_detection_metrics.jpg"),
        bbox_inches="tight",
        pad_inches=0,
    )
    plt.close("all")


def extract_tif_from_bounds(
    reference_image, target_image, output_path, dtype=rasterio.uint16, crs=2154
):
    # Get bounds (intersection)
    with rasterio.open(reference_image) as src:
        bounds = src.bounds
        profile = src.profile
    # Extract image on the same area

    with rasterio.open(target_image) as src:
        window = src.window(*bounds)
        window_data = src.read(window=window)

        # XXX resolution can change (very little) if bounds are not multiple of resolution

        # Update the profile to match the window
        profile.update(
            {
                "height": window_data.shape[1],
                "width": window_data.shape[2],
                "transform": src.window_transform(window),
                "dtype": dtype,
                "count": window_data.shape[0],
                "crs": crs,
            }
        )

    with rasterio.open(output_path, "w", **profile) as dst:
        dst.write(window_data)

    return window_data


def get_vegetation_and_forest_mask(
    forest_mask_gdf,
    classification_raster_path,
    geometry,
    classes_to_keep,
    resolution=1.5,
    resampling_method="bilinear",
):
    bounds = geometry.bounds
    raster_bounds = box(*bounds)
    classification, profile = get_window(
        classification_raster_path,
        geometry=None,
        bounds=bounds,
        resolution=resolution,
        return_profile=True,
        resampling_method=resampling_method,
    )
    classification = classification.squeeze()

    # Create a mask for pixels with value 5 in the classification raster
    classif_mask = classification == classes_to_keep[0]
    if len(classes_to_keep) > 1:
        for aclass in classes_to_keep[1::]:
            classif_mask = classif_mask | (classification == aclass)

    # Clip the geometries to the raster bounds
    clipped_gdf = gpd.clip(forest_mask_gdf, raster_bounds)

    # Rasterize the clipped GeoDataFrame geometries
    if min(classification.shape) == 0:
        print(f"classification is empty {classification.shape=}")
        return np.zeros_like(classif_mask, dtype=bool), profile

    geometries = [(geom, 1) for geom in clipped_gdf.geometry]
    if len(geometries):
        mask_geometries = rasterize(
            geometries,
            out_shape=classification.shape,
            transform=profile["transform"],
            fill=0,
            default_value=1,
            dtype=np.uint8,
        ).astype(bool)
    else:
        mask_geometries = np.zeros_like(classif_mask, dtype=bool)

    # Combine the two masks
    final_mask = classif_mask | mask_geometries
    return final_mask, profile


def get_window(
    image_path,
    geometry=None,
    bounds=None,
    resolution=None,
    return_profile=False,
    resampling_method="bilinear",
):
    """Retrieve a window from an image, within given bounds or within the bounds of a geometry."""
    with rasterio.open(image_path) as src:
        profile = src.profile
        if bounds is None:
            if geometry is not None:
                bounds = geometry.bounds
            else:
                bounds = src.bounds

        window = from_bounds(*bounds, transform=src.transform)
        transform = src.window_transform(window)

        init_resolution = profile["transform"].a

        if (resolution is not None) and (init_resolution != resolution):
            if resampling_method == "max_pooling":
                # Downsample with max pooling
                data = src.read(window=window)
                # Calculate the target shape
                target_width = int((bounds[2] - bounds[0]) / resolution)
                target_height = int((bounds[3] - bounds[1]) / resolution)

                # Resize the data to the target shape, using max pooling within each block
                def max_pooling_resize(image, target_width, target_height):
                    output = np.zeros(
                        (image.shape[0], target_height, target_width)
                    )
                    scale_x = image.shape[2] / target_width
                    scale_y = image.shape[1] / target_height
                    for i in range(target_height):
                        for j in range(target_width):
                            # XXX could replace int by rounding
                            x_start = int(j * scale_x)
                            x_end = int((j + 1) * scale_x)
                            y_start = int(i * scale_y)
                            y_end = int((i + 1) * scale_y)
                            output[:, i, j] = np.max(
                                image[:, y_start:y_end, x_start:x_end]
                            )
                    return output

                data = max_pooling_resize(data, target_width, target_height)
            elif resampling_method == "bilinear":
                scale_factor = init_resolution / resolution
                data = src.read(
                    out_shape=(
                        src.count,
                        int(np.round(window.height * scale_factor, 0)),
                        int(np.round(window.width * scale_factor, 0)),
                    ),
                    resampling=Resampling.bilinear,
                    window=window,
                )
            else:
                raise ValueError(
                    f"{resampling_method} is not a valid resampling method"
                )

            # Update the transform for the new resolution
            transform = Affine(
                resolution,
                transform.b,
                transform.c,
                transform.d,
                -resolution,
                transform.f,
            )
        else:
            data = src.read(window=window)

    if return_profile:
        new_profile = profile.copy()
        if len(data.shape) == 3:
            count = data.shape[0]
            height = data.shape[1]
            width = data.shape[2]
        else:
            count = 1
            height = data.shape[0]
            width = data.shape[1]
        new_profile.update(
            {
                "transform": transform,
                "driver": "GTiff",
                "height": height,
                "width": width,
                "count": count,
            }
        )
        return data, new_profile
    else:
        return data
